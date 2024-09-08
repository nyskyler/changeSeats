from openpyxl import load_workbook
from math import ceil
from random import shuffle
import copy
import time
filePath = "./listOfNames/"
fileName = "6하늘.xlsx"
wb = load_workbook(filePath + fileName)
ws = wb.active

numOfStudents = ws.max_row - 1 
numOfGroups = ceil(numOfStudents/4)
boys = []
girls = []

print(f'numOfStudents: {numOfStudents}')
print(f'numOfGroups: {numOfGroups}')

def getNumOfBoysAndGirls(): # 성별에 근거하여 남학생과 여학생 리스트를 만드는 한편, 남학생과 여학생의 숫자를 반환
  global boys, girls
  numOfBoys = 0
  numOfGirls = 0
  for row in ws.iter_rows(min_row=2, max_row=numOfStudents+1, min_col=4, max_col=5):
    val = row[1].value
    if val == '남성':
      numOfBoys += 1
      boys.append(row[0].value)
    elif val == '여성':
      numOfGirls += 1
      girls.append(row[0].value)
    else:
      continue
  return numOfBoys, numOfGirls

def groupingBoysAndGirls(numOfBoys, numOfGirls, numOfGroups): # 남녀 숫자와 모둠수를 바탕으로 모둠별 남녀 학생 수 이차원 리스트를 반환
  result = []
  for _ in range(numOfGroups):
    result.extend([[0,0]])
  idx = 0
  while numOfBoys > 0 or numOfGirls > 0:
    if numOfBoys > 0 and result[idx][0] + result[idx][1] < 4:
      result[idx][0] += 1
      numOfBoys -= 1
    if numOfGirls > 0 and result[idx][0] + result[idx][1] < 4:
      result[idx][1] += 1
      numOfGirls -= 1
    idx += 1
    idx = idx % numOfGroups
  return result
    
b, g = getNumOfBoysAndGirls()
print(f'남학생: {b}명, 여학생: {g}명')

print(groupingBoysAndGirls(b, g, numOfGroups))  
# print(groupingBoysAndGirls(9, 12, 6))  

def checkCellsEmptyOrNot(): # 비고란이 비었는지 확인
  if any([row[0].value for row in ws.iter_rows(min_row=2, max_row=numOfStudents+1, min_col=6, max_col=6)]):
    return False
  else:
    return True
  
def initialArrangementOfSeats(): # 비고란이 비었을 시 남녀 학생 수 이차원 리스트를 기반으로 무작위 배치
  boysCopied = copy.deepcopy(boys)
  girlsCopied = copy.deepcopy(girls)
  shuffle(boysCopied)
  shuffle(girlsCopied)
  result = []
  temp = groupingBoysAndGirls(b, g, numOfGroups)
  for group in temp:
    sublist = []
    for _ in range(group[0]):
      sublist.append(boysCopied.pop(0))
    for _ in range(group[1]):
      sublist.append(girlsCopied.pop(0))
    result.extend([sublist])
  return result
    
def belongToWhichGroup(name, result): # 주어진 이름의 학생이 몇 모둠에 속하는지 숫자 데이터를 반환
  idx = 0
  for in_list in result:
    if name in in_list:
      return idx
    idx += 1
  return -1

def getListFromFile(): # 엑셀 파일로부터 모둠별 학생 편성 리스트를 만들어 반환
  result = []
  idx = 0
  for _ in range(numOfGroups):
    idx += 1
    in_list = []
    for row in ws.iter_rows(min_row=2, max_row=numOfStudents+1, min_col=4, max_col=6):
      if row[2].value == str(idx):
        in_list.append(row[0].value)
    result.extend([in_list])
  return result

def getSexAndGroupNum(name): # 주어진 이름의 학생의 성별과 모둠 튜플을 반환
  for row in ws.iter_rows(min_row=2, max_row=numOfStudents+1, min_col=4, max_col=6):
    if row[0].value == name:
      history = str(row[2].value)
      return row[1].value, history[len(history)-1]
    else:
      continue
  return -1

def rearrangementOfSeats(groupList): # 비고란이 비어있지 않을 때, 최근 기록을 근거로 자리를 재배치
  orders = [n for n in range(numOfGroups)]
  shuffle(orders)
  shuffle(groupList)
  result = []
  temp = groupingBoysAndGirls(b, g, numOfGroups)
  idx = 0
  for group in groupList:
    for member in group:
      flag = True
      sex, prevGroupNum = getSexAndGroupNum(member)
      sex = 0 if sex == '남성' else 1
      cnt = 0
      while flag:
        time.sleep(0.2)
        if idx > numOfGroups - 1:
          idx = 0
        group_to_check = orders[idx]
        if group_to_check == int(prevGroupNum) - 1:
          idx += 1
          continue
        else:
          remain = temp[group_to_check][sex]
          print(group_to_check, remain)
          if remain > 0:
            temp[group_to_check][sex] -= 1
            result.append((member, group_to_check + 1))
            flag = False
          idx += 1
          cnt += 1
          if cnt > numOfGroups - 2:
            temp[int(prevGroupNum)-1][sex] -= 1
            result.append((member, int(prevGroupNum)))
            flag = False
  return result

def writeResultToFile():
  for row in ws.iter_rows(min_row=2, max_row=numOfStudents+1, min_col=4, max_col=6):
    pos = belongToWhichGroup(row[0].value)
    if pos == -1:
      continue
    else:
      if row[2].value is None:
        row[2].value = str(pos+1)
      else:
        row[2].value = str(row[2].value) + str(pos+1)
  wb.save(filePath + fileName)

def appendResultToFile(result): # rearrangementOfSeats 함수로부터 반환받은 리스트를 전달하여 엑셀 파일에 반영
  for name, group in result:
    for row in ws.iter_rows(min_row=2, max_row=numOfStudents+1, min_col=4, max_col=6):
      if row[0].value == name:
        row[2].value = str(row[2].value) + str(group)
  wb.save(filePath + fileName)

def deleteValues():  # 비고란의 모든 내용을 삭제
  for row in ws.iter_rows(min_row=2, max_row=numOfStudents+1, min_col=6, max_col=6):
    row[0].value = ''
  wb.save(filePath + fileName)

# if checkCellsEmptyOrNot():
#   pass
# else:
#   appendResultToFile(rearrangementOfSeats(getListFromFile()))

# checkCellsEmptyOrNot()
# result = initialArrangementOfSeats()
# # print(result)
# # writeResultToFile()
# # deleteValues()
# print(getListFromFile())
# print(rearrangementOfSeats(getListFromFile()))
deleteValues()