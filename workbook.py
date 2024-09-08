from openpyxl import load_workbook
from math import ceil
from random import shuffle
import copy
import time

class Workbook:
  def __init__(self, filename):
    self.filePath = "./listOfNames/" + filename
    try:
      self.wb = load_workbook(self.filePath)
    except FileNotFoundError:
      print(f"File {filename} not found in ./listOfNames/")
      raise
    self.ws = self.wb.active
    self.numOfStudents = self.ws.max_row - 1
    self.numOfGroups = ceil(self.numOfStudents / 4)
    self.studentNames = []
    self.studentSexes = []
    self.numOfBoys = 0
    self.numOfGirls = 0
    self.lenOfRemarksColumn = 0
    self.index = 1
    self.boys = []
    self.girls = []
    self.getNumOfBoysAndGirls()
    self.listOfNumberTuples = []
    self.groupingBoysAndGirls()
    self.currentArrangement = []
    self.groupNums = []
    self.studentInfoDict = {}
    self.processedResult = []
    self.initializeOrLoadArrangement()

  def getLenOfRemarksColumn(self):
    max = 0
    for row in self.ws.iter_rows(min_row=2, max_row=self.numOfStudents+1, min_col=6, max_col=6):
      max = len(str(row[0].value)) if len(str(row[0].value)) > max else max
    self.lenOfRemarksColumn = max

  def getNumOfBoysAndGirls(self): 
    self.studentNames.clear()
    self.studentSexes.clear()
    self.boys.clear()
    self.girls.clear()
    for row in self.ws.iter_rows(min_row=2, max_row=self.numOfStudents+1, min_col=4, max_col=5):
      self.studentNames.append(row[0].value)
      self.studentSexes.append(row[1].value)
    self.boys = [self.studentNames[idx] for idx in [i for i, sex in enumerate(self.studentSexes) if sex == "남성"]]
    self.girls = [self.studentNames[idx] for idx in [i for i, sex in enumerate(self.studentSexes) if sex == "여성"]]
    self.numOfBoys = len(self.boys)
    self.numOfGirls = len(self.girls)

  def groupingBoysAndGirls(self):
    self.listOfNumberTuples.clear()
    self.listOfNumberTuples = [[0, 0] for _ in range(self.numOfGroups)]
    idx = 0
    numOfBoys = self.numOfBoys
    numOfGirls = self.numOfGirls
    while numOfBoys > 0 or numOfGirls > 0:
      if numOfBoys > 0 and self.listOfNumberTuples[idx][0] + self.listOfNumberTuples[idx][1] < 4:
        self.listOfNumberTuples[idx][0] += 1
        numOfBoys -= 1
      if numOfGirls > 0 and self.listOfNumberTuples[idx][0] + self.listOfNumberTuples[idx][1] < 4:
        self.listOfNumberTuples[idx][1] += 1
        numOfGirls -= 1
      idx += 1
      idx = idx % self.numOfGroups

  def deleteLatestValues(self):
    for row in self.ws.iter_rows(min_row=2, max_row=self.numOfStudents+1, min_col=6, max_col=6):
      row[0].value = row[0].value[:self.lenOfRemarksColumn-1]
    self.wb.save(self.filePath)
    self.initialArrangementOfSeats()

  def initializeOrLoadArrangement(self):
    if all([row[0].value for row in self.ws.iter_rows(min_row=2, max_row=self.numOfStudents+1, min_col=6, max_col=6)]):
      self.getListFromFile()
    else:
      self.deleteLatestValues()
      self.initialArrangementOfSeats()
    self.setStudentInfoDict()
    self.getLenOfRemarksColumn()

  def setStudentInfoDict(self):
    self.studentInfoDict.clear()
    for idx, group in enumerate(self.currentArrangement, start=1):
      for member in group:
        loc = self.studentNames.index(member)
        self.studentInfoDict[member] = (self.studentSexes[loc], idx)
      
  def initialArrangementOfSeats(self):
    boysCopied = copy.deepcopy(self.boys)
    girlsCopied = copy.deepcopy(self.girls)
    shuffle(boysCopied)
    shuffle(girlsCopied)
    self.currentArrangement.clear()
    for group in self.listOfNumberTuples:
      in_list = []
      for _ in range(group[0]):
        in_list.append(boysCopied.pop(0))
      for _ in range(group[1]):
        in_list.append(girlsCopied.pop(0))
      self.currentArrangement.extend([in_list])

  def toPreviousIndex(self):
    self.index += 1
    if self.index > self.lenOfRemarksColumn:
      self.index = self.lenOfRemarksColumn
      return
    self.getListFromFile()
    self.setStudentInfoDict()

  def toNextIndex(self):
    self.index -= 1
    if self.index < 1:
      self.index = 1
      return
    self.getListFromFile()
    self.setStudentInfoDict()

  def getListFromFile(self):
    self.currentArrangement.clear()
    idx = 0
    for _ in range(self.numOfGroups):
      idx += 1
      in_list = []
      for row in self.ws.iter_rows(min_row=2, max_row=self.numOfStudents+1, min_col=4, max_col=6):
        strLen = len(row[2].value)
        if row[2].value[strLen - self.index] == str(idx):
          in_list.append(str(row[0].value))
      self.currentArrangement.extend([in_list])

  def getSexAndGroupNum(self, name):
    return self.studentInfoDict.get(name)
  
  def rearrangementOfSeats(self):
    orders = [n for n in range(self.numOfGroups)]
    listOfNumberTuples = copy.deepcopy(self.listOfNumberTuples)
    currentSeats = copy.deepcopy(self.currentArrangement)
    shuffle(orders)
    shuffle(currentSeats)
    idx = 0
    result = list()
    for group in currentSeats:
      for member in group:
        flag = True
        sex, prevGroupNum = self.getSexAndGroupNum(member)
        sex = 0 if sex == '남성' else 1
        cnt = 0
        while flag:
          time.sleep(0.2)
          if idx > self.numOfGroups - 1:
            idx = 0
          group_to_check = orders[idx]
          if group_to_check == int(prevGroupNum) - 1:
            idx += 1
          else:
            remain = listOfNumberTuples[group_to_check][sex]
            if remain > 0:
              listOfNumberTuples[group_to_check][sex] -= 1
              result.append((member, group_to_check + 1))
              flag = False
            idx += 1
            cnt += 1
            if cnt > self.numOfGroups - 2:
              listOfNumberTuples[int(prevGroupNum)-1][sex] -= 1
              result.append((member, int(prevGroupNum)))
              flag = False
    max_index = max([x[1] for x in result]) - 1
    self.processedResult.clear()
    self.processedResult = [[] for _ in range(max_index + 1)]
    for item in result:
      self.processedResult[item[1] - 1].append(item[0])

  def writeResultToFile(self):
    for idx, group in enumerate(self.processedResult, start=1):
      for name in group:
        for row in self.ws.iter_rows(min_row=2, max_row=self.numOfStudents+1, min_col=4, max_col=6):
          if row[0].value == name:
            row[2].value = str(row[2].value) + str(idx)
            break
    self.wb.save(self.filePath)